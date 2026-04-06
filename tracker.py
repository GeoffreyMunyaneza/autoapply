"""
tracker.py — Manages the Excel job tracking spreadsheet.

Columns:
  ID | Title | Company | Location | Salary | Source | URL | Date Posted
  Match Score | Resume Type | Status | Resume Path | Date Added | Notes
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper import Job

logger = logging.getLogger(__name__)

COLUMNS = [
    "ID",
    "Title",
    "Company",
    "Location",
    "Salary",
    "Source",
    "URL",
    "Date Posted",
    "Match Score",
    "Resume Type",
    "Status",
    "Resume Path",
    "Date Added",
    "Notes",
]

STATUS_COLORS = {
    "Discovered": "FFF9C4",   # yellow
    "Queued":     "C8E6C9",   # green
    "Applied":    "BBDEFB",   # blue
    "Interviewing": "E1BEE7", # purple
    "Offer":      "A5D6A7",   # bright green
    "Rejected":   "FFCDD2",   # red
    "Skipped":    "EEEEEE",   # grey
}

HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _create_workbook(path: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    # Headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = {
        "ID": 20, "Title": 35, "Company": 25, "Location": 20,
        "Salary": 18, "Source": 12, "URL": 40, "Date Posted": 14,
        "Match Score": 12, "Resume Type": 12, "Status": 14,
        "Resume Path": 45, "Date Added": 14, "Notes": 30,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return wb


def _load_or_create(path: str) -> openpyxl.Workbook:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.exists():
        return openpyxl.load_workbook(path)
    return _create_workbook(path)


def load_seen_ids(tracker_path: str) -> set[str]:
    """Return the set of job IDs already in the tracker."""
    p = Path(tracker_path)
    if not p.exists():
        return set()
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb.active
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # ID column
            seen.add(str(row[0]))
    return seen


def add_job(
    tracker_path: str,
    job: Job,
    match_score: float,
    resume_type: str,
    resume_path: str = "",
    status: str = "Discovered",
) -> None:
    """Append a new job row to the tracker."""
    wb = _load_or_create(tracker_path)
    ws = wb.active

    # Find next empty row
    next_row = ws.max_row + 1

    values = [
        job.id,
        job.title,
        job.company,
        job.location,
        job.salary,
        job.source,
        job.url,
        job.date_posted,
        f"{match_score:.0%}",
        resume_type.upper(),
        status,
        resume_path,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        "",  # Notes — user fills in
    ]

    color = STATUS_COLORS.get(status, "FFFFFF")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Make URL clickable
    url_col = COLUMNS.index("URL") + 1
    url_cell = ws.cell(row=next_row, column=url_col)
    if job.url:
        url_cell.hyperlink = job.url
        url_cell.font = Font(color="1565C0", underline="single")

    # Make resume path clickable if it's a local file
    resume_col = COLUMNS.index("Resume Path") + 1
    resume_cell = ws.cell(row=next_row, column=resume_col)
    if resume_path:
        resume_cell.hyperlink = f"file:///{resume_path.replace(chr(92), '/')}"
        resume_cell.font = Font(color="1565C0", underline="single")

    wb.save(tracker_path)


def update_status(tracker_path: str, job_id: str, status: str, notes: str = "") -> bool:
    """Update the status (and optionally notes) for an existing job row."""
    if not Path(tracker_path).exists():
        return False

    wb = openpyxl.load_workbook(tracker_path)
    ws = wb.active

    status_col = COLUMNS.index("Status") + 1
    notes_col = COLUMNS.index("Notes") + 1
    id_col = COLUMNS.index("ID") + 1

    color = STATUS_COLORS.get(status, "FFFFFF")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) == job_id:
            row[status_col - 1].value = status
            row[status_col - 1].fill = fill
            if notes:
                row[notes_col - 1].value = notes
            for cell in row:
                cell.fill = fill
            wb.save(tracker_path)
            return True

    return False
