"""
review.py - Interactive CLI for reviewing tailored resumes.

Usage:
  python review.py
  python review.py --list
  python review.py --approve-all
"""

import argparse
import shutil
import subprocess
import sys
from pathlib import Path

from dotenv import load_dotenv

from services.config import load_runtime_config, resolve_runtime_path


def _find_pending(pending_folder: Path) -> list[Path]:
    """Return pending .docx files sorted by name."""
    if not pending_folder.exists():
        return []
    return sorted(pending_folder.glob("*.docx"))


def _print_diff(diff_file: Path) -> None:
    """Print the diff file with basic ANSI color when supported."""
    if not diff_file.exists():
        print("  (no diff file found)")
        return

    try:
        text = diff_file.read_text(encoding="utf-8")
    except Exception:
        print("  (could not read diff file)")
        return

    supports_color = hasattr(sys.stdout, "isatty") and sys.stdout.isatty()
    red = "\033[31m" if supports_color else ""
    green = "\033[32m" if supports_color else ""
    cyan = "\033[36m" if supports_color else ""
    reset = "\033[0m" if supports_color else ""

    for line in text.splitlines():
        if line.startswith("---") or line.startswith("+++"):
            print(f"{cyan}{line}{reset}")
        elif line.startswith("-"):
            print(f"{red}{line}{reset}")
        elif line.startswith("+"):
            print(f"{green}{line}{reset}")
        else:
            print(line)


def _open_in_word(docx_path: Path) -> None:
    """Try to open the DOCX in Word."""
    try:
        subprocess.Popen(["start", "", str(docx_path)], shell=True)
    except Exception:
        pass


def _update_tracker_by_filename(tracker_path: str, filename: str, status: str) -> None:
    """Update tracker status for the row whose resume path contains filename."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill

        from core.tracker import COLUMNS, STATUS_COLORS

        tracker = Path(tracker_path)
        if not tracker.exists():
            return

        workbook = openpyxl.load_workbook(tracker_path)
        worksheet = workbook.active
        status_col = COLUMNS.index("Status") + 1
        resume_col = COLUMNS.index("Resume Path") + 1
        fill = PatternFill(
            start_color=STATUS_COLORS.get(status, "FFFFFF"),
            end_color=STATUS_COLORS.get(status, "FFFFFF"),
            fill_type="solid",
        )

        for row in worksheet.iter_rows(min_row=2):
            resume_value = str(row[resume_col - 1].value or "")
            if filename.replace(".docx", "") in resume_value:
                row[status_col - 1].value = status
                for cell in row:
                    cell.fill = fill
                workbook.save(tracker_path)
                return
    except Exception:
        pass


def review_one(
    docx_path: Path,
    output_folder: Path,
    tracker_path: str,
    *,
    open_word: bool = False,
) -> str:
    """
    Review a single pending resume.

    Returns one of: approved, rejected, skipped, quit.
    """
    diff_file = docx_path.with_suffix(".diff.txt")

    print("\n" + "=" * 70)
    print(f"  {docx_path.name}")
    print("=" * 70)
    _print_diff(diff_file)

    if open_word:
        _open_in_word(docx_path)

    while True:
        choice = input("\n  [a]pprove  [r]eject  [o]pen in Word  [s]kip  [q]uit -> ").strip().lower()

        if choice in ("a", "approve"):
            destination = output_folder / docx_path.name
            destination.parent.mkdir(parents=True, exist_ok=True)
            counter = 1
            while destination.exists():
                destination = output_folder / f"{docx_path.stem}_{counter}.docx"
                counter += 1

            shutil.move(str(docx_path), str(destination))
            if diff_file.exists():
                diff_file.unlink()
            _update_tracker_by_filename(tracker_path, docx_path.name, "Queued")
            print(f"  Approved -> {destination.name}")
            return "approved"

        if choice in ("r", "reject"):
            docx_path.unlink(missing_ok=True)
            if diff_file.exists():
                diff_file.unlink()
            _update_tracker_by_filename(tracker_path, docx_path.name, "Skipped")
            print("  Rejected; removed from pending")
            return "rejected"

        if choice in ("o", "open"):
            _open_in_word(docx_path)
            continue

        if choice in ("s", "skip"):
            print("  Skipped (stays in pending)")
            return "skipped"

        if choice in ("q", "quit"):
            print("  Quitting review session")
            return "quit"

        print("  Invalid; enter a, r, o, s, or q")


def main() -> None:
    load_dotenv(dotenv_path=resolve_runtime_path(".env"))

    parser = argparse.ArgumentParser(description="AutoApply Resume Review CLI")
    parser.add_argument("--config", default="config.yaml")
    parser.add_argument("--list", action="store_true", help="List pending resumes and exit")
    parser.add_argument("--approve-all", action="store_true", help="Approve all pending resumes without prompting")
    parser.add_argument("--open-word", action="store_true", help="Open each resume in Word during review")
    args = parser.parse_args()

    config = load_runtime_config(args.config)
    output_cfg = config["output"]
    pending_folder = Path(output_cfg.get("pending_folder", "output/pending"))
    resumes_folder = Path(output_cfg["resumes_folder"])
    tracker_path = output_cfg["tracker_file"]
    pending = _find_pending(pending_folder)

    if not pending:
        print("No pending resumes to review.")
        return

    print(f"\nFound {len(pending)} resume(s) pending review in {pending_folder}/")

    if args.list:
        for pending_file in pending:
            diff = pending_file.with_suffix(".diff.txt")
            changed_paragraphs = "?"
            if diff.exists():
                try:
                    changed_paragraphs = sum(
                        1
                        for line in diff.read_text(encoding="utf-8").splitlines()
                        if line.startswith("[Paragraph")
                    )
                except Exception:
                    pass
            print(f"  {pending_file.name}  ({changed_paragraphs} changed paragraphs)")
        return

    if args.approve_all:
        approved = 0
        resumes_folder.mkdir(parents=True, exist_ok=True)
        for pending_file in pending:
            destination = resumes_folder / pending_file.name
            shutil.move(str(pending_file), str(destination))
            diff_file = pending_file.with_suffix(".diff.txt")
            if diff_file.exists():
                diff_file.unlink()
            _update_tracker_by_filename(tracker_path, pending_file.name, "Queued")
            approved += 1
        print(f"Approved all {approved} pending resume(s) -> {resumes_folder}/")
        return

    approved = 0
    rejected = 0
    skipped = 0
    for pending_file in pending:
        result = review_one(
            pending_file,
            resumes_folder,
            tracker_path,
            open_word=args.open_word,
        )
        if result == "approved":
            approved += 1
        elif result == "rejected":
            rejected += 1
        elif result == "skipped":
            skipped += 1
        elif result == "quit":
            break

    print(f"\nReview complete - approved: {approved}, rejected: {rejected}, skipped: {skipped}")
    remaining = len(_find_pending(pending_folder))
    if remaining:
        print(f"{remaining} resume(s) still pending. Run `python review.py` to continue.")


if __name__ == "__main__":
    main()
